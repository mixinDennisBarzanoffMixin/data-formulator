import io
import json
import re
import zipfile
from uuid import UUID

import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from data_formulator.veritly_worker import app


def xlsx() -> bytes:
    book = Workbook()
    rows = book.active
    rows.title = "Transactions"
    for row in [
        ("Customer", "Amount", "Booked"),
        ("Ada", 42, "2026-01-01"),
        ("Lin", 18, "2026-01-02"),
        ("Sam", 27, "2026-01-03"),
        ("Mia", 33, "=C5+6"),
    ]:
        rows.append([None, *row])
    rows["H2"] = "Code"
    rows["I2"] = "Label"
    rows["H3"] = "A"
    rows["I3"] = "Side lookup"
    rows["F20"] = "Unrelated note"
    rows["K100"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    hidden = book.create_sheet("Reference")
    hidden.sheet_state = "hidden"
    hidden["C4"] = "code"
    secret = book.create_sheet("Internal")
    secret.sheet_state = "veryHidden"
    secret["A3"] = "private"
    data = io.BytesIO()
    book.save(data)
    return data.getvalue()


def unsized() -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(xlsx())) as source, zipfile.ZipFile(output, "w") as target:
        for item in source.infolist():
            data = source.read(item)
            if item.filename == "xl/worksheets/sheet1.xml":
                data, count = re.subn(br"<dimension[^>]*/>", b"", data, count=1)
                assert count == 1
            target.writestr(item, data)
    return output.getvalue()


def empty() -> bytes:
    book = Workbook()
    book.active.title = "Empty"
    data = io.BytesIO()
    book.save(data)
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(data.getvalue())) as source, zipfile.ZipFile(output, "w") as target:
        for item in source.infolist():
            content = source.read(item)
            if item.filename == "xl/worksheets/sheet1.xml":
                content, count = re.subn(br"<dimension[^>]*/>", b"", content, count=1)
                assert count == 1
            target.writestr(item, content)
    return output.getvalue()


def styled_blanks() -> bytes:
    book = Workbook()
    sheet = book.active
    sheet.title = "Rows"
    sheet.append(["Customer", "Veritly ID", "Side"])
    sheet.append(["Ada", "", None])
    sheet.append(["", "", "unrelated"])
    sheet.append(["Grace", "", None])
    sheet.append([" ", "", None])
    for cell in [sheet["A3"], sheet["B3"]]:
        cell.fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    data = io.BytesIO()
    book.save(data)
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(data.getvalue())) as source, zipfile.ZipFile(output, "w") as target:
        for item in source.infolist():
            content = source.read(item)
            if item.filename == "xl/worksheets/sheet1.xml":
                for cell in [b"A3", b"B3"]:
                    pattern = rb'(<c r="' + cell + rb'"[^>]*)\s*/>'
                    content, count = re.subn(pattern, rb"\g<1>><is><t></t></is></c>", content, count=1)
                    assert count == 1
            target.writestr(item, content)
    return output.getvalue()


def test_workbook_catalog_returns_only_bounded_sheet_metadata(monkeypatch):
    monkeypatch.setenv("DATA_WORKER_TOKEN", "secret")
    response = app.test_client().post(
        "/workbook",
        data={"file": (io.BytesIO(xlsx()), "rows.xlsx")},
        headers={"x-veritly-service-token": "secret"},
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    assert response.get_json() == {
        "sheets": [
            {
                "name": "Transactions",
                "rows": {"start": 1, "end": 100},
                "columns": {"start": 2, "end": 11},
                "visibility": "visible",
                "regions": [
                    {"header": 1, "start": 2, "end": 5, "left": 2, "right": 4},
                    {"header": 2, "start": 3, "end": 3, "left": 8, "right": 9},
                ],
            },
            {
                "name": "Reference",
                "rows": {"start": 4, "end": 4},
                "columns": {"start": 3, "end": 3},
                "visibility": "hidden",
                "regions": [],
            },
            {
                "name": "Internal",
                "rows": {"start": 3, "end": 3},
                "columns": {"start": 1, "end": 1},
                "visibility": "veryHidden",
                "regions": [],
            },
        ]
    }


def test_workbook_catalog_streams_dimensions_when_the_xlsx_omits_them(monkeypatch):
    monkeypatch.setenv("DATA_WORKER_TOKEN", "secret")
    response = app.test_client().post(
        "/workbook",
        data={"file": (io.BytesIO(unsized()), "unsized.xlsx")},
        headers={"x-veritly-service-token": "secret"},
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    assert response.get_json()["sheets"][0] == {
        "name": "Transactions",
        "rows": {"start": 1, "end": 100},
        "columns": {"start": 2, "end": 11},
        "visibility": "visible",
        "regions": [
            {"header": 1, "start": 2, "end": 5, "left": 2, "right": 4},
            {"header": 2, "start": 3, "end": 3, "left": 8, "right": 9},
        ],
    }


def test_workbook_catalog_defaults_an_unsized_empty_sheet_to_a1(monkeypatch):
    monkeypatch.setenv("DATA_WORKER_TOKEN", "secret")
    response = app.test_client().post(
        "/workbook",
        data={"file": (io.BytesIO(empty()), "empty.xlsx")},
        headers={"x-veritly-service-token": "secret"},
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    assert response.get_json() == {
        "sheets": [{
            "name": "Empty",
            "rows": {"start": 1, "end": 1},
            "columns": {"start": 1, "end": 1},
            "visibility": "visible",
            "regions": [],
        }]
    }


def test_workbook_catalog_requires_service_authentication(monkeypatch):
    monkeypatch.setenv("DATA_WORKER_TOKEN", "secret")
    response = app.test_client().post(
        "/workbook",
        data={"file": (io.BytesIO(xlsx()), "rows.xlsx")},
        content_type="multipart/form-data",
    )
    assert response.status_code == 401


def test_normalize_and_locate_omit_explicit_empty_selected_rows(monkeypatch):
    monkeypatch.setenv("DATA_WORKER_TOKEN", "secret")
    client = app.test_client()
    data = styled_blanks()
    selection = json.dumps({"sheet": "Rows", "header": 1, "start": 2, "end": 5, "left": 1, "right": 2})
    headers = {"x-veritly-service-token": "secret"}

    normalized = client.post(
        "/normalize",
        data={"selection": selection, "file": (io.BytesIO(data), "rows.xlsx")},
        headers=headers,
        content_type="multipart/form-data",
    )
    assert normalized.status_code == 200
    rows = pq.read_table(io.BytesIO(normalized.data)).to_pylist()
    assert rows == [
        {"Customer": "Ada", "Veritly ID": None},
        {"Customer": "Grace", "Veritly ID": None},
        {"Customer": " ", "Veritly ID": None},
    ]

    located = client.post(
        "/locate",
        data={
            "selection": selection,
            "identity": "Veritly ID",
            "file": (io.BytesIO(data), "rows.xlsx"),
        },
        headers=headers,
        content_type="multipart/form-data",
    )
    assert located.status_code == 200
    assert [json.loads(line) for line in located.get_data(as_text=True).splitlines()] == [
        {"$veritly": {"columns": ["Customer", "Veritly ID"]}},
        {"row": 1, "id": None},
        {"row": 3, "id": None},
        {"row": 4, "id": None},
    ]

    commands = [
        {"id": "source", "after": [], "kind": "source", "output": "raw"},
        {
            "id": "key",
            "after": ["source"],
            "kind": "key",
            "input": "raw",
            "output": "keyed",
            "key": {"strategy": "generated", "name": "Veritly ID"},
        },
        {
            "id": "final",
            "after": ["key"],
            "kind": "output",
            "input": "keyed",
            "schema": "rows_12345678",
            "table": "rows",
            "class": "entity",
            "keys": ["Veritly ID"],
            "owners": {"Veritly ID": "workbook"},
        },
    ]
    executed = client.post(
        "/execute",
        data={
            "recipe": json.dumps({"inputs": ["raw"], "commands": commands, "output": "final"}),
            "file": (io.BytesIO(normalized.data), "raw.parquet"),
        },
        headers={**headers, "accept": "application/x-ndjson"},
        content_type="multipart/form-data",
    )
    assert executed.status_code == 200
    output = [json.loads(line) for line in executed.get_data(as_text=True).splitlines()][1:]
    ids = [UUID(row["Veritly ID"]) for row in output]
    assert len(ids) == 3
    assert len(set(ids)) == 3
    assert [row["Customer"] for row in output] == ["Ada", "Grace", " "]


def test_workbook_catalog_audits_archive_paths_before_opening(monkeypatch):
    data = io.BytesIO()
    with zipfile.ZipFile(data, "w") as archive:
        archive.writestr("../outside.xml", "unsafe")
    monkeypatch.setenv("DATA_WORKER_TOKEN", "secret")
    response = app.test_client().post(
        "/workbook",
        data={"file": (io.BytesIO(data.getvalue()), "unsafe.xlsx")},
        headers={"x-veritly-service-token": "secret"},
        content_type="multipart/form-data",
    )
    assert response.status_code == 400
    assert response.get_json() == {"error": "XLSX archive contains an unsafe path"}
