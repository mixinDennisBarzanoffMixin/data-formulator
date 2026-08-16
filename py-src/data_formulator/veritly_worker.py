"""Stateless XLSX profiling and normalization worker for Veritly project data."""

from __future__ import annotations

import base64
import hashlib
import json
import os
import shutil
import tempfile
import zipfile
import xml.etree.ElementTree as etree
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path, PurePosixPath
from typing import Iterator
from uuid import UUID

import duckdb
import pyarrow as pa
import pyarrow.parquet as pq
from flask import Flask, Response, after_this_request, jsonify, request, send_file, stream_with_context
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from data_formulator.veritly_recipe import RecipeError, execute as execute_recipe, rewrite_metadata


MAX_BYTES = 100 * 1024 * 1024
MAX_EXPANDED = 2 * 1024 * 1024 * 1024
MAX_RATIO = 200
MAX_CELLS = 20_000_000
MAX_ROWS = 1_048_576
MAX_COLUMNS = 16_384
MAX_SHEETS = 1_024
MAX_REGIONS = 32
BATCH = 10_000

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = MAX_BYTES


def authorized() -> bool:
    token = os.environ.get("DATA_WORKER_TOKEN", "").strip()
    if not token:
        raise RuntimeError("DATA_WORKER_TOKEN is required")
    return request.headers.get("x-veritly-service-token") == token


def selection() -> dict[str, object]:
    raw = request.form.get("selection")
    if not raw:
        raise ValueError("selection is required")
    value = json.loads(raw)
    if not isinstance(value, dict):
        raise ValueError("selection must be an object")
    return value


def upload() -> Path:
    file = request.files.get("file")
    if not file:
        raise ValueError("XLSX file is required")
    fd, name = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    path = Path(name)
    file.save(path)
    if path.stat().st_size > MAX_BYTES:
        path.unlink(missing_ok=True)
        raise ValueError("XLSX exceeds 100 MiB")
    try:
        audit(path)
    except Exception:
        path.unlink(missing_ok=True)
        raise
    return path


def recipe() -> dict[str, object]:
    raw = request.form.get("recipe")
    if not raw:
        raise RecipeError("invalid_input", "recipe is required")
    try:
        value = json.loads(raw)
    except json.JSONDecodeError as error:
        raise RecipeError("invalid_input", "recipe must be valid JSON") from error
    if not isinstance(value, dict):
        raise RecipeError("invalid_input", "recipe must be an object")
    return value


def parquet_inputs(root: Path, value: dict[str, object]) -> dict[str, Path]:
    names = value.get("inputs")
    files = request.files.getlist("file")
    if not isinstance(names, list) or not names or len(names) != len(files):
        raise RecipeError("invalid_input", "recipe inputs must match the staged files")
    output: dict[str, Path] = {}
    total = 0
    for offset, item in enumerate(names):
        if not isinstance(item, str) or not item.strip() or len(item) > 256 or item in output:
            raise RecipeError("invalid_input", "recipe input identifiers must be unique")
        target = root / f"input-{offset}.parquet"
        files[offset].save(target)
        total += target.stat().st_size
        if total > MAX_BYTES:
            raise RecipeError("invalid_input", "Parquet inputs exceed 100 MiB")
        output[item] = target
    return output


def digest(value: dict[str, object]) -> str:
    raw = json.dumps(value, separators=(",", ":")).encode()
    return hashlib.sha256(raw).hexdigest()


def scalar(value: object) -> object:
    if isinstance(value, (datetime, date, time, UUID, Decimal)):
        return str(value)
    if isinstance(value, bytes):
        return base64.b64encode(value).decode()
    return value


def ndjson(path: Path, manifest: dict[str, object], root: Path) -> Iterator[str]:
    try:
        yield json.dumps({"$veritly": manifest}, separators=(",", ":")) + "\n"
        source = pq.ParquetFile(path)
        for batch in source.iter_batches(batch_size=BATCH):
            for row in batch.to_pylist():
                yield json.dumps({key: scalar(value) for key, value in row.items()}, separators=(",", ":")) + "\n"
    finally:
        shutil.rmtree(root, ignore_errors=True)


def audit(path: Path) -> None:
    try:
        archive = zipfile.ZipFile(path)
    except zipfile.BadZipFile as error:
        raise ValueError("Malformed XLSX archive") from error
    with archive:
        total = 0
        for item in archive.infolist():
            name = PurePosixPath(item.filename)
            if name.is_absolute() or ".." in name.parts:
                raise ValueError("XLSX archive contains an unsafe path")
            total += item.file_size
            if total > MAX_EXPANDED:
                raise ValueError("XLSX expanded content exceeds 2 GiB")
            if item.compress_size and item.file_size / item.compress_size > MAX_RATIO:
                raise ValueError("XLSX archive has an unsafe compression ratio")


def index(value: object) -> int:
    if isinstance(value, int) and value > 0:
        return value
    if isinstance(value, str) and value.isalpha():
        return column_index_from_string(value.upper())
    raise ValueError("Columns must be positive indexes or Excel letters")


def catalog(path: Path) -> dict[str, object]:
    book = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    archive = zipfile.ZipFile(path)
    try:
        if len(book.worksheets) > MAX_SHEETS:
            raise ValueError("XLSX exceeds the worksheet limit")
        sheets = []
        for sheet in book.worksheets:
            bounds, found = inspect(archive, sheet)
            left, start, right, end = bounds
            if not all(isinstance(item, int) for item in [left, start, right, end]):
                raise ValueError("Worksheet has no bounded used range")
            if start < 1 or end > MAX_ROWS or end < start:
                raise ValueError("Worksheet used rows exceed Excel bounds")
            if left < 1 or right > MAX_COLUMNS or right < left:
                raise ValueError("Worksheet used columns exceed Excel bounds")
            if sheet.sheet_state not in {"visible", "hidden", "veryHidden"}:
                raise ValueError("Worksheet visibility is invalid")
            sheets.append({
                "name": sheet.title,
                "rows": {"start": start, "end": end},
                "columns": {"start": left, "end": right},
                "visibility": sheet.sheet_state,
                "regions": found,
            })
        return {"sheets": sheets}
    finally:
        archive.close()
        book.close()


def inspect(archive: zipfile.ZipFile, sheet) -> tuple[tuple[int, int, int, int], list[dict[str, int]]]:
    source = getattr(sheet, "_worksheet_path", None)
    if not isinstance(source, str) or not source:
        raise ValueError("Worksheet XML path is unavailable")
    active: list[dict[str, int]] = []
    found: list[dict[str, int]] = []
    bounds: list[int] | None = None
    with archive.open(source) as body:
        for _, element in etree.iterparse(body, events=("end",)):
            if element.tag.rsplit("}", 1)[-1] != "row":
                continue
            number = row_number(element)
            raw, columns = row_columns(element, number)
            if raw and bounds is None:
                bounds = [min(raw), number, max(raw), number]
            elif raw and bounds is not None:
                bounds = [
                    min(bounds[0], min(raw)),
                    min(bounds[1], number),
                    max(bounds[2], max(raw)),
                    max(bounds[3], number),
                ]
            next_active: list[dict[str, int]] = []
            for item in active:
                hits = sum(item["left"] <= column <= item["right"] for column in columns)
                if hits:
                    item["end"] = number
                    item["rows"] += 1
                    item["cells"] += hits
                    next_active.append(item)
                    continue
                if number - item["end"] <= 2:
                    next_active.append(item)
                    continue
                found.append(item)
            spans = runs(columns)
            spans = [span for span in spans if not any(
                item["left"] == span[0] and item["right"] == span[1] and item["end"] == number
                for item in next_active
            )]
            next_active.extend({
                "header": number,
                "start": number + 1,
                "end": number,
                "left": left,
                "right": right,
                "rows": 1,
                "cells": right - left + 1,
            } for left, right in spans if number < MAX_ROWS)
            active = next_active
            element.clear()
    found.extend(active)
    candidates = [item for item in found if item["rows"] > 1 and item["end"] >= item["start"]]
    candidates.sort(key=lambda item: (
        -(item["rows"] - 1),
        -item["cells"],
        -(item["right"] - item["left"] + 1),
        item["header"],
        item["left"],
    ))
    selected: list[dict[str, int]] = []
    for item in candidates:
        region = {name: item[name] for name in ["header", "start", "end", "left", "right"]}
        if region in selected:
            continue
        if any(
            region["header"] >= current["header"]
            and region["end"] <= current["end"]
            and region["left"] >= current["left"]
            and region["right"] <= current["right"]
            for current in selected
        ):
            continue
        selected.append(region)
        if len(selected) == MAX_REGIONS:
            break
    if bounds is None:
        return (1, 1, 1, 1), selected
    return (bounds[0], bounds[1], bounds[2], bounds[3]), selected


def row_number(row) -> int:
    raw = row.attrib.get("r")
    if not isinstance(raw, str) or not raw.isdigit():
        raise ValueError("Worksheet row number is invalid")
    number = int(raw)
    if number < 1 or number > MAX_ROWS:
        raise ValueError("Worksheet row exceeds Excel bounds")
    return number


def row_columns(row, number: int) -> tuple[list[int], list[int]]:
    raw: list[int] = []
    columns = []
    for cell in row:
        if cell.tag.rsplit("}", 1)[-1] != "c":
            continue
        reference = cell.attrib.get("r")
        if not isinstance(reference, str):
            raise ValueError("Worksheet cell reference is invalid")
        letters = reference.rstrip("0123456789")
        digits = reference[len(letters):]
        if not letters or not digits.isdigit() or int(digits) != number:
            raise ValueError("Worksheet cell reference is invalid")
        column = column_index_from_string(letters)
        if column < 1 or column > MAX_COLUMNS:
            raise ValueError("Worksheet column exceeds Excel bounds")
        raw.append(column)
        if populated(cell):
            columns.append(column)
    return sorted(set(raw)), sorted(set(columns))


def populated(cell) -> bool:
    return any(
        child.tag.rsplit("}", 1)[-1] == "f"
        or child.tag.rsplit("}", 1)[-1] in {"v", "t"} and child.text not in {None, ""}
        for child in cell.iter()
        if child is not cell
    )


def runs(columns: list[int]) -> list[tuple[int, int]]:
    if not columns:
        return []
    output: list[tuple[int, int]] = []
    start = columns[0]
    end = start
    for column in columns[1:]:
        if column == end + 1:
            end = column
            continue
        output.append((start, end))
        start = column
        end = column
    output.append((start, end))
    return output


def text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def config(value: dict[str, object]) -> tuple[str, int, int, int, list[int]]:
    sheet = value.get("sheet")
    if not isinstance(sheet, str) or not sheet:
        raise ValueError("sheet is required")
    header = value.get("header")
    start = value.get("start")
    end = value.get("end")
    if not all(isinstance(item, int) and item > 0 for item in [header, start, end]):
        raise ValueError("Row bounds must be positive integers")
    if int(start) <= int(header) or int(end) < int(start):
        raise ValueError("Invalid row bounds")
    columns = value.get("columns", [])
    if not isinstance(columns, list):
        raise ValueError("columns must be an array")
    selected = [index(item) for item in columns]
    if not selected:
        left = value.get("left")
        right = value.get("right")
        if not isinstance(left, int) or not isinstance(right, int) or left < 1 or right < left:
            raise ValueError("A bounded column range is required")
        selected = list(range(left, right + 1))
    return sheet, int(header), int(start), int(end), selected


def names(row: tuple[object, ...], columns: list[int]) -> tuple[list[int], list[str]]:
    selected = columns
    output: list[str] = []
    used: dict[str, int] = {}
    for column in selected:
        base = text(row[column - 1] if column <= len(row) else None) or f"Column {column}"
        used[base] = used.get(base, 0) + 1
        output.append(base if used[base] == 1 else f"{base} ({used[base]})")
    return selected, output


def rows(path: Path, value: dict[str, object]) -> tuple[list[str], Iterator[list[str | None]]]:
    sheet, header, start, end, columns = config(value)
    book = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    if sheet not in book.sheetnames:
        book.close()
        raise ValueError("Worksheet not found")
    tab = book[sheet]
    if (end - start + 1) * len(columns) > MAX_CELLS:
        book.close()
        raise ValueError("Selected range exceeds the cell limit")
    header_row = next(tab.iter_rows(min_row=header, max_row=header, values_only=True))
    selected, fields = names(header_row, columns)

    def stream() -> Iterator[list[str | None]]:
        try:
            for row in tab.iter_rows(min_row=start, max_row=end, values_only=True):
                values = [text(row[column - 1] if column <= len(row) else None) for column in selected]
                if any(value is not None for value in values):
                    yield values
        finally:
            book.close()

    return fields, stream()


def parquet(source: Path, value: dict[str, object]) -> tuple[Path, int, list[str]]:
    fields, stream = rows(source, value)
    schema = pa.schema([pa.field(field, pa.string()) for field in fields])
    fd, name = tempfile.mkstemp(suffix=".parquet")
    os.close(fd)
    target = Path(name)
    writer = pq.ParquetWriter(target, schema, compression="zstd")
    count = 0
    batch: list[list[str | None]] = []
    try:
        for row in stream:
            batch.append(row)
            if len(batch) < BATCH:
                continue
            writer.write_table(
                pa.Table.from_arrays(
                    [pa.array(column, type=pa.string()) for column in zip(*batch)],
                    schema=schema,
                )
            )
            count += len(batch)
            batch = []
        if batch:
            writer.write_table(
                pa.Table.from_arrays(
                    [pa.array(column, type=pa.string()) for column in zip(*batch)],
                    schema=schema,
                )
            )
            count += len(batch)
    finally:
        writer.close()
    return target, count, fields


def quote(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def stats(path: Path, fields: list[str], count: int) -> dict[str, object]:
    db = duckdb.connect(":memory:")
    db.read_parquet(str(path)).create_view("data")
    columns = []
    for field in fields:
        name = quote(field)
        values = db.execute(
            f"select count(*) filter (where {name} is null or trim({name}) = ''), "
            f"count(distinct {name}), count(*) filter (where try_cast({name} as double) is not null), "
            f"count(*) filter (where try_cast({name} as timestamp) is not null) from data"
        ).fetchone()
        if not values:
            raise RuntimeError("Profile query returned no result")
        columns.append({
            "name": field,
            "nulls": values[0],
            "unique": values[1],
            "numeric": values[2],
            "dates": values[3],
        })
    duplicates = 0
    if fields:
        group = ", ".join(quote(field) for field in fields)
        duplicates = db.execute(f"select coalesce(sum(n - 1), 0) from (select count(*) n from data group by {group})").fetchone()[0]
    db.close()
    return {"rows": count, "duplicates": duplicates, "columns": columns}


def lineage(path: Path, value: dict[str, object], fields: list[str]) -> dict[str, object]:
    sheet, _, start, end, columns = config(value)
    source = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    cached = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    if sheet not in source.sheetnames or sheet not in cached.sheetnames:
        source.close()
        cached.close()
        raise ValueError("Worksheet not found")
    output = {field: {"formulas": 0, "stale": 0, "expressions": []} for field in fields}
    raw = source[sheet].iter_rows(min_row=start, max_row=end, values_only=True)
    values = cached[sheet].iter_rows(min_row=start, max_row=end, values_only=True)
    try:
        for left, right in zip(raw, values):
            for offset, column in enumerate(columns):
                formula = left[column - 1] if column <= len(left) else None
                if not isinstance(formula, str) or not formula.startswith("="):
                    continue
                output[fields[offset]]["formulas"] += 1
                if formula not in output[fields[offset]]["expressions"] and len(output[fields[offset]]["expressions"]) < 20:
                    output[fields[offset]]["expressions"].append(formula)
                value = right[column - 1] if column <= len(right) else None
                if value is None:
                    output[fields[offset]]["stale"] += 1
    finally:
        source.close()
        cached.close()
    return output


@app.get("/livez")
def livez():
    return jsonify({"ok": True})


@app.get("/metrics")
def metrics():
    return Response(
        "# HELP veritly_data_worker_up Whether the data preparation worker is running.\n"
        "# TYPE veritly_data_worker_up gauge\n"
        "veritly_data_worker_up 1\n",
        mimetype="text/plain",
    )


@app.get("/readyz")
def readyz():
    if not os.environ.get("DATA_WORKER_TOKEN", "").strip():
        return jsonify({"ok": False}), 503
    return jsonify({"ok": True})


@app.post("/profile")
def profile():
    if not authorized():
        return jsonify({"error": "unauthorized"}), 401
    source = upload()
    target = None
    try:
        value = selection()
        target, count, fields = parquet(source, value)
        result = stats(target, fields, count)
        formulas = lineage(source, value, fields)
        for column in result["columns"]:
            column.update(formulas[column["name"]])
        result["formulaIssues"] = sum(item["stale"] for item in formulas.values())
        return jsonify(result)
    finally:
        source.unlink(missing_ok=True)
        if target:
            target.unlink(missing_ok=True)


@app.post("/workbook")
def workbook():
    if not authorized():
        return jsonify({"error": "unauthorized"}), 401
    source = upload()
    try:
        return jsonify(catalog(source))
    finally:
        source.unlink(missing_ok=True)


@app.post("/preview")
def preview():
    if not authorized():
        return jsonify({"error": "unauthorized"}), 401
    source = upload()
    target = None
    try:
        value = selection()
        page = int(request.form.get("page", "0"))
        size = min(int(request.form.get("size", "100")), 100)
        if page < 0 or size < 1:
            raise ValueError("Invalid page")
        target, count, fields = parquet(source, value)
        db = duckdb.connect(":memory:")
        table = db.execute("select * from read_parquet(?) limit ? offset ?", [str(target), size, page * size])
        output = [dict(zip(fields, row)) for row in table.fetchall()]
        db.close()
        return jsonify({"columns": fields, "rows": output, "total": count})
    finally:
        source.unlink(missing_ok=True)
        if target:
            target.unlink(missing_ok=True)


@app.post("/locate")
def locate():
    if not authorized():
        return jsonify({"error": "unauthorized"}), 401
    source = upload()
    book = None
    try:
        value = selection()
        identity = request.form.get("identity", "").strip()
        if not identity:
            raise ValueError("identity is required")
        sheet, header, start, end, columns = config(value)
        book = load_workbook(source, read_only=True, data_only=True, keep_links=False)
        if sheet not in book.sheetnames:
            raise ValueError("Worksheet not found")
        tab = book[sheet]
        header_row = next(tab.iter_rows(min_row=header, max_row=header, values_only=True))
        selected, fields = names(header_row, columns)
        matches = [offset for offset, field in enumerate(fields) if field == identity]
        if len(matches) != 1:
            raise ValueError("identity column must occur exactly once in the selected range")
        identity_column = selected[matches[0]]
    except Exception:
        if book:
            book.close()
        source.unlink(missing_ok=True)
        raise

    def generate():
        try:
            yield json.dumps({"$veritly": {"columns": fields}}, separators=(",", ":")) + "\n"
            for number, row in enumerate(tab.iter_rows(min_row=start, max_row=end, values_only=True), start=start):
                values = [row[column - 1] if column <= len(row) else None for column in selected]
                if not any(item is not None for item in values):
                    continue
                yield json.dumps({
                    "row": number - 1,
                    "id": text(row[identity_column - 1] if identity_column <= len(row) else None),
                }, separators=(",", ":")) + "\n"
        finally:
            book.close()
            source.unlink(missing_ok=True)

    return Response(stream_with_context(generate()), mimetype="application/x-ndjson")


@app.post("/normalize")
def normalize():
    if not authorized():
        return jsonify({"error": "unauthorized"}), 401
    source = upload()
    value = selection()
    target, _, fields = parquet(source, value)
    formulas = lineage(source, value, fields)
    rewrite_metadata(target, b"veritly.lineage", {
        field: {
            "owner": "formula" if formulas[field]["formulas"] else "shared",
            **({"formula": formulas[field]["expressions"][0]} if formulas[field]["expressions"] else {}),
            "stale": formulas[field]["stale"] > 0,
            "staleCount": formulas[field]["stale"],
        }
        for field in fields
    })
    source.unlink(missing_ok=True)

    @after_this_request
    def cleanup(response):
        target.unlink(missing_ok=True)
        return response

    return send_file(target, mimetype="application/vnd.apache.parquet", as_attachment=True, download_name="dataset.parquet")


@app.post("/execute")
def execute():
    if not authorized():
        return jsonify({"error": "unauthorized"}), 401
    root = Path(tempfile.mkdtemp(prefix="veritly-recipe-"))
    try:
        value = recipe()
        commands = value.get("commands")
        if not isinstance(commands, list):
            raise RecipeError("invalid_input", "recipe commands must be an array")
        output = value.get("output")
        if not isinstance(output, str) or not output:
            raise RecipeError("invalid_input", "recipe output is required")
        lineage = value.get("lineage")
        if lineage is not None and not isinstance(lineage, dict):
            raise RecipeError("invalid_input", "recipe lineage must be an object")
        rows = value.get("rows", 1_000_000)
        if not isinstance(rows, int):
            raise RecipeError("invalid_input", "recipe rows must be an integer")
        result = execute_recipe(
            root,
            parquet_inputs(root, value),
            commands,
            output,
            lineage,
            value.get("writeback") is True,
            rows,
        )
        headers = {
            "X-Veritly-Recipe-Sha256": digest(result.manifest),
            "X-Content-Type-Options": "nosniff",
        }
        if "application/x-ndjson" in request.headers.get("accept", ""):
            return Response(
                stream_with_context(ndjson(result.path, result.manifest, root)),
                mimetype="application/x-ndjson",
                headers=headers,
            )

        @after_this_request
        def cleanup_recipe(response):
            shutil.rmtree(root, ignore_errors=True)
            return response

        response = send_file(
            result.path,
            mimetype="application/vnd.apache.parquet",
            as_attachment=True,
            download_name="dataset.parquet",
            conditional=False,
            etag=False,
        )
        response.headers.update(headers)
        return response
    except Exception:
        shutil.rmtree(root, ignore_errors=True)
        raise


@app.errorhandler(ValueError)
def invalid(error: ValueError):
    if isinstance(error, RecipeError):
        return jsonify({"error": str(error), "code": error.code, "command": error.command}), 400
    return jsonify({"error": str(error)}), 400


def run() -> None:
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", "8080")), debug=False)


if __name__ == "__main__":
    run()
